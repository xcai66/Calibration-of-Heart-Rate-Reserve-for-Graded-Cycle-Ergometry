#!/usr/bin/env python3
"""Build the auditable CycHRR-T results workbook with openpyxl."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1E5A8A")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEAF4")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BLUE_BOLD = Font(color="173A53", bold=True)
THIN = Side(style="thin", color="D9E0E5")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def fit_columns(sheet, max_width: int = 48) -> None:
    for column in range(1, sheet.max_column + 1):
        width = 0
        for cell in sheet[get_column_letter(column)]:
            if cell.value is not None:
                width = max(width, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), max_width)


def style_table(sheet, header_row: int = 1) -> None:
    sheet.freeze_panes = f"A{header_row + 1}"
    for cell in sheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row):
        for cell in row:
            cell.border = ALL_BORDERS
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    fit_columns(sheet)


def add_csv_sheet(book: Workbook, name: str, path: Path) -> None:
    sheet = book.create_sheet(name)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            sheet.append(row)
    style_table(sheet)


def build_readme(book: Workbook) -> None:
    sheet = book.active
    sheet.title = "README"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "CycHRR-T analysis results workbook"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 30
    rows = [
        ("Field", "Description"),
        ("Purpose", "Auditable tables supporting endpoint-preserving HRR calibration in graded cycle ergometry"),
        ("Locked model", "Normalized quadratic tail; tau = 0.90; kappa = 5.75"),
        ("Primary target", "Oxygen-uptake reserve (VO2R)"),
        ("Primary unit", "Complete graded-test file; participant identity across files is unavailable"),
        ("External unit", "ACTES participant"),
        ("Development data", "Earlier 70% of tests within sport; grouped five-fold cross-validation"),
        ("Temporal validation", "Latest-date 30% of cycling test files; 84 analysis units"),
        ("External validation", "ACTES PhysioNet; 18 participants"),
        ("Interpretation", "Negative MAE difference favors CycHRR-T; nonlinear superiority was not general"),
        ("Repository", "https://github.com/xcai66/Calibration-of-Heart-Rate-Reserve-for-Graded-Cycle-Ergometry"),
        ("Author ORCID", "https://orcid.org/0009-0002-3662-4539"),
    ]
    for row_index, row in enumerate(rows, start=3):
        sheet.cell(row_index, 1, row[0])
        sheet.cell(row_index, 2, row[1])
        for cell in sheet[row_index][:2]:
            cell.border = ALL_BORDERS
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 1).font = Font(bold=True)
    for cell in sheet[3][:2]:
        cell.fill = SUBHEADER_FILL
        cell.font = BLUE_BOLD
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 82


def build_key_results(book: Workbook) -> None:
    sheet = book.create_sheet("Key Results")
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Locked validation and practical magnitude"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    rows = [
        ["Dataset", "Target", "n", "Raw HRR MAE", "CycHRR-T MAE", "Absolute change", "Change (points)", "Relative change", "Exact 10% agreement"],
        ["Temporal holdout", "VO2R", 84, 0.0616978544, 0.0509500397, None, None, None, 0.5970],
        ["Temporal holdout", "Load fraction", 84, 0.0461176353, 0.0398548278, None, None, None, 0.6042],
        ["ACTES external", "VO2R", 18, 0.0716972959, 0.0579428112, None, None, None, 0.5112],
        ["ACTES external", "Power fraction", 18, 0.1032121191, 0.0807314840, None, None, None, 0.4221],
    ]
    for r_index, row in enumerate(rows, start=3):
        for c_index, value in enumerate(row, start=1):
            sheet.cell(r_index, c_index, value)
    for r_index in range(4, 8):
        sheet.cell(r_index, 6, f"=E{r_index}-D{r_index}")
        sheet.cell(r_index, 7, f"=F{r_index}*100")
        sheet.cell(r_index, 8, f"=F{r_index}/D{r_index}")
        for c_index in range(4, 7):
            sheet.cell(r_index, c_index).number_format = "0.0000"
        sheet.cell(r_index, 7).number_format = "0.0000"
        sheet.cell(r_index, 8).number_format = "0.0%"
        sheet.cell(r_index, 9).number_format = "0.0%"
    style_table(sheet, header_row=3)


def main() -> None:
    args = parse_args()
    root = args.root.resolve()
    output = args.output.resolve()
    book = Workbook()
    build_readme(book)
    build_key_results(book)
    csv_sheets = [
        ("Extraction Audit", "02_data/derived/graded_tests_extraction_audit.csv"),
        ("Sport Selection", "04_results/tables/sport_direction_development_comparison.csv"),
        ("Cycling Dev CV", "04_results/tables/development_grouped_cv_summary.csv"),
        ("Locked Validation", "04_results/tables/locked_validation_summary.csv"),
        ("Locked Per Unit", "04_results/tables/locked_validation_per_unit.csv"),
        ("Strong Comparators", "04_results/tables/strong_comparator_summary.csv"),
        ("Strong Per Unit", "04_results/tables/strong_comparator_per_unit.csv"),
        ("Endpoint Exclusion", "04_results/tables/endpoint_exclusion_summary.csv"),
        ("Endpoint Per Unit", "04_results/tables/endpoint_exclusion_per_unit.csv"),
        ("Intensity Bands", "04_results/tables/intensity_band_summary.csv"),
        ("Intensity Per Unit", "04_results/tables/intensity_band_per_unit.csv"),
        ("10pct Agreement", "04_results/tables/ten_percent_bin_agreement_summary.csv"),
        ("10pct Comparisons", "04_results/tables/ten_percent_bin_agreement_comparisons.csv"),
        ("10pct Per Unit", "04_results/tables/ten_percent_bin_agreement_per_unit.csv"),
        ("Timing Sensitivity", "04_results/tables/actes_lag_sensitivity.csv"),
        ("Parameter Sensitivity", "04_results/tables/parameter_sensitivity.csv"),
        ("Anchor Sensitivity", "04_results/tables/anchor_sensitivity.csv"),
        ("Predefined Splits", "04_results/tables/predefined_test_splits.csv"),
    ]
    for name, relative_path in csv_sheets:
        add_csv_sheet(book, name, root / relative_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)
    print(f"Workbook written: {output}")


if __name__ == "__main__":
    main()
